import requests
import json
import os
import openpyxl

# Function to read data from Excel
def read_excel(file):
    try:
        workbook = openpyxl.load_workbook(file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
 
    # Select the active worksheet (or create one if it doesn't exist)
    sheet = workbook.active
   
    # Dictionary to hold the data from the sheet
    data = {"hosts": []}
 
    # Loop through rows starting from row 2 (assuming row 1 contains headers)
    for i in range(2, sheet.max_row + 1):
       
        # Extract the specific data from columns
        hostName_input = sheet.cell(row=i, column=1).value  # Column 1: Device Name
        monitoringIP_input = sheet.cell(row=i, column=2).value  # Column 2: IP Address
        groupName_input = sheet.cell(row=i, column=3).value  # Column 3: Group
        templateName_input = sheet.cell(row=i, column=4).value  # Column 4: Template
        snmpVersion_input = sheet.cell(row=i, column=5).value  # Column 5: SNMP Version
        proxyName_input = sheet.cell(row=i, column=6).value  # Column 6: Proxy
       
        # Check if the row is empty (no device name)
        if hostName_input is None or monitoringIP_input is None or groupName_input is None or templateName_input is None or snmpVersion_input is None or proxyName_input is None:
            print("Error while processing", hostName_input)
            break
        else:
            print ("Processing", hostName_input)
 
        # Add the row data to the main list
        data["hosts"].append({
            "hostName": hostName_input,
            "monitoringIP": monitoringIP_input,
            "groupName": groupName_input,
            "templateName": templateName_input,
            "snmpVersion": snmpVersion_input,
            "proxyName": proxyName_input
        })
 
    return data

# Get Group ID from Group Name
def get_group_id(group_name, api_key, url, headers):
    payload = {
        "jsonrpc": "2.0",
        "method": "hostgroup.get",
        "params": {
            "output": "extend",
            "filter": {
                "name": group_name
            }
        },
        "auth": api_key,
        "id": 1
    }
    response = requests.post(url, json=payload, headers=headers)
    result = response.json().get("result")
    if result and len(result) > 0:
        return result[0]["groupid"]
    else:
        return None

# Get Template ID from Template Name
def get_template_id(template_name, api_key, url, headers):
    payload = {
        "jsonrpc": "2.0",
        "method": "template.get",
        "params": {
            "output": "extend",
            "filter": {
                "name": template_name
            }
        },
        "auth": api_key,
        "id": 1
    }
    response = requests.post(url, json=payload, headers=headers)
    result = response.json().get("result")
    if result and len(result) > 0:
        return result[0]["templateid"]
    else:
        return None

# Create host
def create_host(current_host, api_key, url, headers, group_id, template_id, snmp_version, communityString, snmpUsername, snmpPassword, monitorStatus):
    payload = {
        "jsonrpc": "2.0",
        "method": "host.create",
        "params": {
            "host": current_host["hostName"],
            "interfaces": [
                {
                    "type": 2,
                    "main": 1,
                    "useip": 1,
                    "ip": current_host["monitoringIP"],
                    "dns": "",
                    "port": "161",
                }
            ],
            "groups": [{"groupid": group_id}],
            "templates": [{"templateid": template_id}],
            "inventory_mode": 1
        },
        "id": 1,
        "auth": api_key,
    }

    if snmp_version == 2:
        payload["params"]["interfaces"][0]["details"] = {
            "version": 2,
            "bulk": 1,
            "community": "{$SNMP_COMMUNITY}",
        }
        payload["params"]["macros"] = {
            "macro": "{$SNMP_COMMUNITY}",
            "value": communityString,
        }
    elif snmp_version == 3:
        payload["params"]["interfaces"][0]["details"] = {
            "version": 3,
            "bulk": 1,
            "securityname": snmpUsername,
            "contextname": snmpUsername,
            "securitylevel": 1,
            "authprotocol": 5,
            "authpassphrase": snmpPassword,
        }

    response = requests.post(url, json=payload, headers=headers)
    return response.json()

def main():
    api_key = ""
    url = "http://IP-ADDRESS/zabbix/api_jsonrpc.php"
    headers = {"Content-Type": "application/json"}
    snmpUsername = ""
    snmpPassword = ""
    communityString = ""
    file = os.path.join(os.getcwd(), "input.xlsx")
    data = read_excel(file)
    
    for current_host in data["hosts"]:
        group_id = get_group_id(current_host["groupName"], api_key, url, headers)
        template_id = get_template_id(current_host["templateName"], api_key, url, headers)
        snmp_version = (current_host["snmpVersion"])
        if group_id==None:
            print(f"Error: Group ID not found for {current_host['hostName']}")
            continue
        elif template_id==None:
            print(f"Error: Template ID not found for {current_host['hostName']}")
            continue
        elif snmp_version!=2 and snmp_version!=3:
            print(f"Error: SNMP Version not found for {current_host['hostName']}")
            continue

        result = create_host(current_host, api_key, url, headers, group_id, template_id, snmp_version, communityString, snmpUsername, snmpPassword, monitorStatus)
        if "result" in result: 
            print("SUCCESS: Host " + current_host['hostName'] + " ID is " + result['result']['hostids'][0])
        elif "error" in result:
         print("FAILURE: " + result['error']['message'] + " " + result['error']['data'])

if __name__=="__main__":
    main()
